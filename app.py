"""
Dashboard Bagi Hasil Teknisi (aplikasi berdiri sendiri)
=======================================================
Menghitung omzet jasa per teknisi beserta bagi hasilnya, dengan:
  - tarif per kata kunci pada NAMA BARANG yang bisa diubah manual
  - periode penggajian memakai cutoff tanggal 24 s/d 23
  - perbandingan terhadap skema flat (seluruh omzet jasa x satu tarif)

Jalankan:
    pip install -r requirements.txt
    streamlit run app.py
"""
import io
import re
from datetime import date
from pathlib import Path

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

st.set_page_config(page_title="Bagi Hasil Teknisi", layout="wide", page_icon="🧰")

DEFAULT_SALES_PATH = Path(__file__).parent / "data" / "penjualan.csv.gz"

BULAN_NAMES = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli',
               'Agustus', 'September', 'Oktober', 'November', 'Desember']

PALETTE = ['#1f3864', '#2e9bd6', '#16a34a', '#e0921f', '#c9392f',
           '#7c3aed', '#0f8a82', '#a855f7', '#3f8ac9', '#d1478d']

# --- aturan bagi hasil (nilai awal; bisa diubah dari dashboard) --------------
KATA_KUNCI_TARIF = ['INTERFACE', 'NORMAL', 'MATI TOTAL', 'PROMO']
TARIF_AWAL = {'Interface': 20.0, 'Normal': 30.0, 'Mati Total': 32.0, 'Promo': 60.0}
TARIF_DEFAULT_AWAL = 30.0
TARIF_PEMBANDING_AWAL = 30.0
LABEL_LAINNYA = 'Lainnya'

st.markdown("""
<style>
  .kpi-wrap{display:grid;grid-template-columns:repeat(6,1fr);gap:14px;margin-bottom:6px;}
  @media(max-width:1200px){.kpi-wrap{grid-template-columns:repeat(3,1fr);}}
  .kpi{border-radius:16px;padding:16px 16px 18px;color:#fff;min-height:112px;
       box-shadow:0 8px 20px rgba(30,20,60,.14);}
  .kpi .label{font-size:11px;font-weight:800;text-transform:uppercase;letter-spacing:.04em;opacity:.92;}
  .kpi .value{font-size:21px;font-weight:800;margin-top:10px;line-height:1.15;}
  .kpi .foot{font-size:11px;margin-top:6px;opacity:.9;}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Utilitas
# ---------------------------------------------------------------------------
def rp(v, singkat=True):
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "-"
    neg, v = v < 0, abs(float(v))
    if singkat:
        if v >= 1_000_000_000:
            s = f"Rp {v/1_000_000_000:,.2f} M"
        elif v >= 1_000_000:
            s = f"Rp {v/1_000_000:,.1f} jt"
        elif v >= 1_000:
            s = f"Rp {v/1_000:,.0f} rb"
        else:
            s = f"Rp {v:,.0f}"
    else:
        s = f"Rp {v:,.0f}"
    s = s.replace(",", "#").replace(".", ",").replace("#", ".")
    return ("-" + s) if neg else s


def kpi_html(cards):
    cells = []
    for c in cards:
        cells.append(f"""
        <div class="kpi" style="background:{c['grad']}">
          <div class="label">{c['label']}</div>
          <div class="value">{c['value']}</div>
          <div class="foot">{c.get('sub','&nbsp;')}</div>
        </div>""")
    return f'<div class="kpi-wrap">{"".join(cells)}</div>'


def cocok_kata_kunci(nama_barang):
    s = str(nama_barang).upper()
    return [k for k in KATA_KUNCI_TARIF if k in s]


def pilih_label_tarif(kw_str, urutan):
    if not kw_str:
        return LABEL_LAINNYA
    cocok = kw_str.split('|')
    for k in urutan:
        if k in cocok:
            return k.title()
    return cocok[0].title()


def periode_gaji(bulan_gaji: int, tahun_gaji: int):
    """Gaji bulan M dihitung dari 24 bulan (M-2) s/d 23 bulan (M-1).

    Contoh: gaji Mei 2026 -> 24 Maret 2026 s/d 23 April 2026.
    """
    m_akhir, th_akhir = bulan_gaji - 1, tahun_gaji
    if m_akhir < 1:
        m_akhir += 12
        th_akhir -= 1
    m_awal, th_awal = m_akhir - 1, th_akhir
    if m_awal < 1:
        m_awal += 12
        th_awal -= 1
    return pd.Timestamp(th_awal, m_awal, 24), pd.Timestamp(th_akhir, m_akhir, 23)


def label_periode(bulan_gaji, tahun_gaji):
    a, b = periode_gaji(bulan_gaji, tahun_gaji)
    return (f"Gaji {BULAN_NAMES[bulan_gaji]} {tahun_gaji}  "
            f"({a.day} {BULAN_NAMES[a.month]} – {b.day} {BULAN_NAMES[b.month]} {b.year})")


def daftar_periode_gaji(tgl_min, tgl_max):
    hasil = []
    if pd.isna(tgl_min) or pd.isna(tgl_max):
        return hasil
    y, m = tgl_min.year, tgl_min.month
    for _ in range(120):
        a, b = periode_gaji(m, y)
        if a > tgl_max:
            break
        if b >= tgl_min:
            hasil.append((y, m))
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return hasil


SALES_REQUIRED = ['TGL FAKTUR', 'NO FAKTUR', 'KATEGORI BARANG', 'NAMA BARANG',
                  'QTY', 'TOTAL HARGA']          # CABANG boleh datang dari nama berkas
KOLOM_DIPAKAI = SALES_REQUIRED + ['CABANG', 'NAMA TEKNISI', 'NAMA TEKNISI (FINAL)']
# NO FAKTUR hanya unik DI DALAM satu cabang (nomor MF-FP.xxxx dipakai ulang di
# cabang lain), jadi kunci duplikat wajib menyertakan CABANG. Baris kembar di
# dalam satu berkas tetap dipertahankan — yang dibuang hanya kiriman ulang.
KUNCI_DUPLIKAT = ['CABANG', 'NO FAKTUR', 'NAMA BARANG', 'QTY', 'TOTAL HARGA',
                  'TGL FAKTUR']
LABEL_TANPA_CABANG = '(TANPA CABANG)'

# Daftar cabang resmi. Nama berkas kiriman cabang biasanya terpotong
# (mis. "001mflashklende") sehingga dicocokkan lewat awalan ke daftar ini.
CABANG_KANONIK = [
    'BINTARA', 'CEGER', 'CIBINONG', 'CIBUBUR', 'CIKAMPEK', 'CILANGKAP', 'CINERE',
    'CONDET', 'DRAMAGA', 'JATIBENING', 'JATIMULYA', 'JATIWARINGIN', 'KARAWANG',
    'KLENDER', 'PEJATEN', 'RADJIMAN', 'SAWANGAN', 'WARBONG',
]
# nama lain -> nama cabang resmi
ALIAS_CABANG_AWAL = {'TELUK JAMBE': 'KARAWANG', 'TELUKJAMBE': 'KARAWANG'}

# potongan kata yang dibuang saat menebak nama cabang
NOISE_NAMA = {
    'RINCIAN', 'PENJUALAN', 'PENJUALANAN', 'DATA', 'LAPORAN', 'LAP', 'REKAP', 'JASA',
    'SALES', 'FAKTUR', 'INVOICE', 'PERIODE', 'BULAN', 'TAHUN', 'CABANG', 'CAB',
    'BRANCH', 'TOKO', 'FIX', 'FINAL', 'REVISI', 'REV', 'UPDATE', 'BARU', 'NEW',
    'COPY', 'SALINAN', 'XLSX', 'XLS', 'CSV', 'GZ', 'FILE', 'BAGI', 'HASIL',
    'TEKNISI', 'OK', 'SHEET', 'LEMBAR', 'TABEL', 'TABLE',
    'JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI', 'JULI', 'AGUSTUS',
    'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER',
    'JAN', 'FEB', 'MAR', 'APR', 'JUN', 'JUL', 'AGU', 'AGS', 'SEP', 'OKT', 'NOV', 'DES',
}
AWALAN_BUANG = ('MFLASH', 'MFLSH', 'MFLAS')      # kode perusahaan di nama berkas


def _huruf(s) -> str:
    return re.sub(r'[^A-Z]', '', str(s).upper())


def token_cabang(nama: str) -> str:
    """Sisa nama berkas/sheet setelah angka, kode, dan kata umum dibuang.

    'rincian_faktur_penjualan_001mflashklende_260818094705.xlsx' -> 'KLENDE'
    'Rincian Faktur Penjualan' -> '' (semuanya kata umum)
    """
    s = re.sub(r'\.(xlsx|xlsm|xls|csv|gz|txt)$', '', str(nama), flags=re.I)
    s = re.sub(r'\.(csv|xlsx)$', '', s, flags=re.I)              # untuk nama .csv.gz
    s = re.sub(r'\b[0-9a-f]{8,}\b', ' ', s, flags=re.I)          # buang kode acak/uuid
    s = re.sub(r'[\_\-\.\(\)\[\]#]+', ' ', s)
    s = re.sub(r'\d+', ' ', s)
    tok = []
    for t in s.upper().split():
        for aw in AWALAN_BUANG:
            if t.startswith(aw) and len(t) > len(aw):
                t = t[len(aw):]
        if len(t) > 1 and t not in NOISE_NAMA:
            tok.append(t)
    return ' '.join(tok).strip()


def cocokkan_cabang(tok: str, kanonik, alias):
    """Cocokkan token ke daftar cabang resmi. -> (nama_cabang, keterangan)."""
    t = _huruf(tok)
    if not t:
        return '', ''
    for a, v in alias.items():
        A = _huruf(a)
        if A and (A.startswith(t) or t.startswith(A)):
            return v, f'alias {a}'
    kandidat = [c for c in kanonik
                if _huruf(c).startswith(t) or t.startswith(_huruf(c))]
    if len(kandidat) == 1:
        return kandidat[0], ''
    if len(kandidat) > 1:
        return tok.upper(), 'ambigu: ' + '/'.join(kandidat)
    return tok.upper(), 'di luar daftar'


def _cabang_dari_kolom(d: pd.DataFrame) -> str:
    """Kalau kolom CABANG terisi seragam, pakai itu. '' kalau kosong/beragam."""
    if 'CABANG' not in d.columns:
        return ''
    v = d['CABANG'].dropna().astype(str).str.strip()
    v = v[(v != '') & (~v.str.upper().isin(['NAN', 'NONE']))]
    if v.empty:
        return ''
    return '' if v.nunique() > 1 else v.iloc[0]     # beragam -> biarkan apa adanya


def _potongan_berkas(nama_berkas: str, isi: bytes):
    """Pecah satu berkas jadi (DataFrame, nama_bagian). Bagian = sheet untuk xlsx."""
    low = str(nama_berkas).lower()
    if low.endswith('.gz'):
        yield pd.read_csv(io.BytesIO(isi), compression='gzip'), ''
    elif low.endswith(('.csv', '.txt')):
        yield pd.read_csv(io.BytesIO(isi)), ''
    else:
        xls = None
        for mesin in ('calamine', 'openpyxl'):      # calamine jauh lebih cepat
            try:
                xls = pd.ExcelFile(io.BytesIO(isi), engine=mesin)
                break
            except Exception:                        # noqa: BLE001, PERF203
                continue
        if xls is None:
            xls = pd.ExcelFile(io.BytesIO(isi), engine='openpyxl')
        for sheet in xls.sheet_names:
            d = xls.parse(sheet)
            if not d.empty:
                yield d, sheet


@st.cache_data(show_spinner="Membaca berkas penjualan...")
def baca_mentah(items: tuple, kanonik: tuple, alias_items: tuple):
    """Gabung banyak berkas jadi satu tabel mentah + catatan asal tiap potongan.

    items: tuple of (nama_berkas, bytes). Nama cabang dicari berurutan:
    kolom CABANG -> nama sheet -> nama berkas -> kosong (diisi manual di sidebar).
    """
    alias = dict(alias_items)
    frames, catatan, gagal = [], [], []
    for nama_berkas, isi in items:
        try:
            potongan = list(_potongan_berkas(nama_berkas, isi))
        except Exception as e:                                   # noqa: BLE001
            gagal.append(f"{nama_berkas}: {e}")
            continue
        if not potongan:
            gagal.append(f"{nama_berkas}: tidak ada baris data")
            continue
        for d, bagian in potongan:
            kurang = [c for c in SALES_REQUIRED if c not in d.columns]
            if kurang:
                gagal.append(f"{nama_berkas}"
                             + (f" [{bagian}]" if bagian else "")
                             + ": kolom tidak ditemukan — " + ", ".join(kurang))
                continue
            d = d[[c for c in KOLOM_DIPAKAI if c in d.columns]].copy()   # hemat memori

            cab, asal, ket = _cabang_dari_kolom(d), 'kolom CABANG', ''
            if not cab and bagian:
                tok = token_cabang(bagian)
                if tok:
                    cab, ket = cocokkan_cabang(tok, kanonik, alias)
                    asal = 'nama sheet'
            if not cab:
                tok = token_cabang(nama_berkas)
                if tok:
                    cab, ket = cocokkan_cabang(tok, kanonik, alias)
                    asal = 'nama berkas'
            beragam = ('CABANG' in d.columns) and bool(d['CABANG'].notna().any())
            if not cab:
                asal, ket = ('kolom CABANG', 'beragam per baris') if beragam \
                    else ('—', 'perlu diisi manual')

            if cab:
                d['CABANG'] = cab
            elif not beragam:
                d['CABANG'] = pd.NA
            d['__BERKAS__'] = nama_berkas
            d['__BAGIAN__'] = bagian
            frames.append(d)
            catatan.append({
                'Berkas': nama_berkas, 'Bagian': bagian or '—',
                'Cabang': cab if cab else ('(beragam)' if beragam else LABEL_TANPA_CABANG),
                'Dari': asal, 'Catatan': ket, 'Baris': len(d)})
    if not frames:
        return pd.DataFrame(), pd.DataFrame(catatan), gagal
    return (pd.concat(frames, ignore_index=True, sort=False),
            pd.DataFrame(catatan), gagal)


def bersihkan(df: pd.DataFrame) -> pd.DataFrame:
    """Normalisasi kolom + saring baris kategori JASA (dipakai semua sumber data)."""
    if df.empty:
        return df
    df = df.copy()
    for c in ['QTY', 'TOTAL HARGA']:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    df['TGL'] = pd.to_datetime(df['TGL FAKTUR'], errors='coerce')
    df['KATEGORI'] = df['KATEGORI BARANG'].astype(str).str.strip().str.upper()
    df['BARANG'] = df['NAMA BARANG'].astype(str).str.strip()
    df['CABANG'] = (df['CABANG'].astype(str).str.strip()
                    .replace({'': LABEL_TANPA_CABANG, 'nan': LABEL_TANPA_CABANG,
                              'NaN': LABEL_TANPA_CABANG, 'None': LABEL_TANPA_CABANG})
                    .fillna(LABEL_TANPA_CABANG))

    fin = df['NAMA TEKNISI (FINAL)'] if 'NAMA TEKNISI (FINAL)' in df.columns \
        else pd.Series(index=df.index, dtype=object)
    asli = df['NAMA TEKNISI'] if 'NAMA TEKNISI' in df.columns \
        else pd.Series(index=df.index, dtype=object)
    tek = fin.fillna(asli).astype(str).str.strip().str.upper()
    tek = tek.replace({'NAN': '', 'NONE': '', '<NA>': ''}).fillna('')
    df['TEKNISI'] = tek
    df.loc[df['TEKNISI'] == '', 'TEKNISI'] = 'TIDAK ADA TEKNISI'

    df = df[df['KATEGORI'] == 'JASA'].copy()
    df['KW_MATCH'] = df['BARANG'].map(lambda s: '|'.join(cocok_kata_kunci(s)))
    return df


@st.cache_data(show_spinner="Membaca data penjualan...")
def load_sales(file_bytes: bytes, source_kind: str) -> pd.DataFrame:
    """Loader berkas tunggal (dipakai untuk data bawaan repo)."""
    nama = {'csv_gz': 'penjualan.csv.gz', 'csv': 'penjualan.csv'}.get(source_kind, 'penjualan.xlsx')
    mentah, _, gagal = baca_mentah(((nama, file_bytes),), tuple(CABANG_KANONIK),
                                   tuple(ALIAS_CABANG_AWAL.items()))
    if mentah.empty:
        raise ValueError(gagal[0] if gagal else "tidak ada baris data")
    return bersihkan(mentah)


# ---------------------------------------------------------------------------
# Sidebar: sumber data (mendukung banyak berkas — satu kiriman per cabang)
# ---------------------------------------------------------------------------
st.sidebar.title("📁 Sumber Data")
ups = st.sidebar.file_uploader(
    "Upload data penjualan — bisa banyak berkas sekaligus",
    type=['xlsx', 'xlsm', 'gz', 'csv'], accept_multiple_files=True,
    key='uploader_cabang',
    help="Kirim satu berkas per cabang (boleh 25 sekaligus), atau satu berkas gabungan. "
         "Kalau kosong, dipakai berkas bawaan data/penjualan.csv.gz.")

buang_duplikat = st.sidebar.checkbox(
    "Buang kiriman ulang (duplikat antar berkas)", value=True, key='opsi_dedup',
    help="Kalau satu cabang mengirim berkas dua kali, baris yang sama persis "
         "(cabang, no faktur, barang, qty, total, tanggal) hanya dihitung sekali — "
         "yang dipakai berkas terakhir. Baris kembar di dalam satu berkas tetap utuh.")

with st.sidebar.expander("🏷️ Daftar & alias cabang", expanded=False):
    st.caption("Nama berkas kiriman cabang sering terpotong (mis. `001mflashklende`). "
               "Potongan itu dicocokkan ke daftar di bawah lewat awalan nama.")
    teks_kanonik = st.text_area(
        "Daftar cabang resmi (satu per baris)",
        value=st.session_state.get('teks_kanonik', "\n".join(CABANG_KANONIK)),
        height=150, key='teks_kanonik')
    teks_alias = st.text_area(
        "Alias — format `nama lain = CABANG RESMI`",
        value=st.session_state.get(
            'teks_alias', "\n".join(f"{k} = {v}" for k, v in ALIAS_CABANG_AWAL.items())),
        height=90, key='teks_alias',
        help="Contoh: TELUK JAMBE = KARAWANG")

kanonik = tuple(dict.fromkeys(
    b.strip().upper() for b in teks_kanonik.splitlines() if b.strip()))
alias_items = tuple(
    (a.strip().upper(), b.strip().upper())
    for a, _, b in (baris.partition('=') for baris in teks_alias.splitlines())
    if a.strip() and b.strip())

jasa_all = pd.DataFrame()
catatan_berkas = pd.DataFrame()
try:
    if ups:
        items = tuple((u.name, u.getvalue()) for u in ups)
        mentah, catatan_berkas, gagal = baca_mentah(items, kanonik, alias_items)
        mentah = mentah.copy()

        # --- koreksi manual untuk potongan yang cabangnya tak terdeteksi ---
        if not mentah.empty and not catatan_berkas.empty:
            perlu = catatan_berkas['Cabang'] == LABEL_TANPA_CABANG
            if perlu.any():
                st.sidebar.warning(f"{int(perlu.sum())} berkas belum ketahuan cabangnya — "
                                   "isi manual di bawah.")
                with st.sidebar.form('form_cabang'):
                    isian = {}
                    for _, r in catatan_berkas[perlu].iterrows():
                        label = r['Berkas'] + (f" [{r['Bagian']}]" if r['Bagian'] != '—' else '')
                        isian[(r['Berkas'], r['Bagian'])] = st.text_input(
                            f"Cabang untuk {label}", key=f"cab_{label}")
                    if st.form_submit_button("Terapkan nama cabang"):
                        st.session_state['peta_cabang'] = {
                            k: v.strip().upper() for k, v in isian.items() if v.strip()}
                        st.rerun()
            for (berkas, bagian), cab in st.session_state.get('peta_cabang', {}).items():
                m = (mentah['__BERKAS__'] == berkas) & \
                    (mentah['__BAGIAN__'] == ('' if bagian == '—' else bagian))
                mentah.loc[m, 'CABANG'] = cab
                mc = (catatan_berkas['Berkas'] == berkas) & (catatan_berkas['Bagian'] == bagian)
                catatan_berkas.loc[mc, 'Cabang'] = cab
                catatan_berkas.loc[mc, 'Dari'] = 'isian manual'
                catatan_berkas.loc[mc, 'Catatan'] = ''

        n_dup = 0
        if not mentah.empty and buang_duplikat and \
                all(c in mentah.columns for c in KUNCI_DUPLIKAT):
            sebelum = len(mentah)
            # nomor urut kemunculan di dalam satu berkas -> baris kembar yang memang
            # ada di berkas aslinya tidak ikut terbuang, hanya kiriman ulang
            mentah['__N__'] = mentah.groupby(KUNCI_DUPLIKAT + ['__BERKAS__'],
                                             dropna=False).cumcount()
            mentah = mentah.drop_duplicates(subset=KUNCI_DUPLIKAT + ['__N__'],
                                            keep='last').drop(columns='__N__')
            n_dup = sebelum - len(mentah)

        jasa_all = bersihkan(mentah)
        if gagal:
            st.sidebar.error("Berkas dilewati:\n\n- " + "\n- ".join(gagal))
        if not jasa_all.empty:
            st.sidebar.success(
                f"{len(ups)} berkas · {jasa_all['CABANG'].nunique()} cabang · "
                f"{len(jasa_all):,} baris jasa"
                + (f" · {n_dup:,} baris duplikat dibuang" if n_dup else ""))
    elif DEFAULT_SALES_PATH.exists():
        jasa_all = load_sales(DEFAULT_SALES_PATH.read_bytes(), 'csv_gz')
        st.sidebar.info("Memakai data bawaan repo.")
except Exception as e:  # noqa: BLE001
    st.sidebar.error(f"Data tidak terbaca: {e}")

if not catatan_berkas.empty:
    with st.sidebar.expander(f"📄 Rincian {len(catatan_berkas)} berkas terbaca", expanded=True):
        st.dataframe(catatan_berkas, hide_index=True, use_container_width=True)
        ragu = catatan_berkas[catatan_berkas['Catatan'].astype(str).str.startswith(
            ('ambigu', 'di luar daftar', 'perlu'))]
        if len(ragu):
            st.warning("Periksa berkas ini — nama cabangnya belum pasti: "
                       + ", ".join(ragu['Berkas'].astype(str)))
        ganda = (catatan_berkas[catatan_berkas['Cabang'] != LABEL_TANPA_CABANG]
                 .groupby('Cabang')['Berkas'].nunique())
        ganda = ganda[ganda > 1]
        if len(ganda):
            st.info("Cabang dengan lebih dari satu berkas: " + ", ".join(ganda.index))

st.title("🧰 Bagi Hasil Teknisi")

if jasa_all.empty:
    st.info(
        "Data belum tersedia. Letakkan file **data/penjualan.csv.gz** di folder aplikasi, "
        "atau upload file penjualan lewat panel kiri.\n\n"
        "Format yang dibutuhkan: data faktur penjualan dengan kolom TGL FAKTUR, NO FAKTUR, "
        "KATEGORI BARANG, NAMA BARANG, NAMA TEKNISI (FINAL), QTY, TOTAL HARGA, dan CABANG "
        "(atau satu sheet per cabang bila berupa .xlsx)."
    )
    st.stop()

st.caption(
    f"{len(jasa_all):,} baris jasa · {jasa_all['CABANG'].nunique()} cabang · "
    f"{jasa_all.loc[jasa_all['TEKNISI'] != 'TIDAK ADA TEKNISI', 'TEKNISI'].nunique()} teknisi · "
    f"data {jasa_all['TGL'].min():%d %b %Y} – {jasa_all['TGL'].max():%d %b %Y}"
)

# ---------------------------------------------------------------------------
# Pengaturan tarif
# ---------------------------------------------------------------------------
with st.expander("⚙️ Pengaturan Tarif Bagi Hasil — klik untuk mengubah", expanded=False):
    st.caption("Ubah angka sesuai kebijakan; seluruh perhitungan langsung menyesuaikan.")
    c1, c2, c3, c4 = st.columns(4)
    tarif_input = {}
    with c1:
        tarif_input['Interface'] = st.number_input(
            "Interface (%)", 0.0, 100.0, TARIF_AWAL['Interface'], 1.0, key='t_int')
    with c2:
        tarif_input['Normal'] = st.number_input(
            "Normal (%)", 0.0, 100.0, TARIF_AWAL['Normal'], 1.0, key='t_nor')
    with c3:
        tarif_input['Mati Total'] = st.number_input(
            "Mati Total (%)", 0.0, 100.0, TARIF_AWAL['Mati Total'], 1.0, key='t_mat')
    with c4:
        tarif_input['Promo'] = st.number_input(
            "Promo (%)", 0.0, 100.0, TARIF_AWAL['Promo'], 1.0, key='t_pro')

    c5, c6, c7 = st.columns([1, 1, 1.6])
    with c5:
        tarif_lain = st.number_input(
            "Tanpa kata kunci (%)", 0.0, 100.0, TARIF_DEFAULT_AWAL, 1.0, key='t_lain',
            help="Untuk item seperti JASA REPAIR / JASA BATERAI yang tidak mengandung kata kunci.")
    with c6:
        tarif_flat = st.number_input(
            "Tarif pembanding (%)", 0.0, 100.0, TARIF_PEMBANDING_AWAL, 1.0, key='t_flat',
            help="Skema pembanding: seluruh omzet jasa dikali tarif ini.")
    with c7:
        prioritas = st.selectbox(
            "Kalau satu nama mengandung 2 kata kunci, yang menang:",
            ['Normal', 'Promo', 'Mati Total', 'Interface'], index=0, key='t_prio')

    if st.button("↩️ Kembalikan ke tarif awal", key='t_reset'):
        for k, v in [('t_int', 'Interface'), ('t_nor', 'Normal'),
                     ('t_mat', 'Mati Total'), ('t_pro', 'Promo')]:
            st.session_state[k] = TARIF_AWAL[v]
        st.session_state['t_lain'] = TARIF_DEFAULT_AWAL
        st.session_state['t_flat'] = TARIF_PEMBANDING_AWAL
        st.session_state['t_prio'] = 'Normal'
        st.rerun()

urutan = [prioritas.upper()] + [k for k in KATA_KUNCI_TARIF if k != prioritas.upper()]
peta_tarif = {k: v / 100.0 for k, v in tarif_input.items()}
peta_tarif[LABEL_LAINNYA] = tarif_lain / 100.0

jasa_all = jasa_all.copy()
jasa_all['TARIF_LABEL'] = jasa_all['KW_MATCH'].map(lambda s: pilih_label_tarif(s, urutan))
jasa_all['TARIF'] = jasa_all['TARIF_LABEL'].map(peta_tarif).fillna(0.0)
jasa_all['BAGI_HASIL'] = jasa_all['TOTAL HARGA'] * jasa_all['TARIF']
jasa_all['FLAT'] = jasa_all['TOTAL HARGA'] * (tarif_flat / 100.0)

st.caption(
    "**Tarif aktif:** " +
    " · ".join(f"{k} {v:.0f}%" for k, v in tarif_input.items()) +
    f" · Lainnya {tarif_lain:.0f}% · pembanding flat {tarif_flat:.0f}%"
    f" · prioritas bentrok: {prioritas}"
)

# ---------------------------------------------------------------------------
# Filter periode & cabang
# ---------------------------------------------------------------------------
fa, fb, fc = st.columns([2.2, 1.4, 1])
periode_list = daftar_periode_gaji(jasa_all['TGL'].min(), jasa_all['TGL'].max())
opsi = ['Semua Periode'] + periode_list
with fa:
    pilih = st.selectbox(
        "Periode penggajian (cutoff tanggal 24 s/d 23)", opsi,
        index=len(opsi) - 1 if len(opsi) > 1 else 0,
        format_func=lambda x: ("Semua Periode (tanpa cutoff)" if isinstance(x, str)
                               else label_periode(x[1], x[0])),
        key='f_periode')
with fb:
    cab_opts = ['Semua Cabang'] + sorted(jasa_all['CABANG'].dropna().unique().tolist())
    f_cabang = st.selectbox("Cabang", cab_opts, key='f_cabang')
with fc:
    sembunyikan = st.checkbox("Sembunyikan baris tanpa nama teknisi", value=False,
                              key='f_hide')

jasa = jasa_all
if f_cabang != 'Semua Cabang':
    jasa = jasa[jasa['CABANG'] == f_cabang]
if isinstance(pilih, str):
    periode_txt = "Seluruh periode data (tanpa cutoff)"
    tag_file = "semua-periode"
else:
    a, b = periode_gaji(pilih[1], pilih[0])
    jasa = jasa[(jasa['TGL'] >= a) & (jasa['TGL'] <= b)]
    periode_txt = (f"{a.day} {BULAN_NAMES[a.month]} {a.year} – "
                   f"{b.day} {BULAN_NAMES[b.month]} {b.year}")
    tag_file = f"gaji-{pilih[0]}-{pilih[1]:02d}"

jasa_tampil = jasa[jasa['TEKNISI'] != 'TIDAK ADA TEKNISI'] if sembunyikan else jasa

st.markdown(f"**Rentang dihitung:** {periode_txt}"
            + (f" · cabang **{f_cabang}**" if f_cabang != 'Semua Cabang' else ""))

if jasa.empty:
    st.warning("Tidak ada transaksi jasa pada periode/cabang tersebut.")
    st.stop()

# ---------------------------------------------------------------------------
# KPI
# ---------------------------------------------------------------------------
omzet = jasa['TOTAL HARGA'].sum()
bh = jasa['BAGI_HASIL'].sum()
fl = jasa['FLAT'].sum()
selisih = bh - fl
n_tek = jasa.loc[jasa['TEKNISI'] != 'TIDAK ADA TEKNISI', 'TEKNISI'].nunique()
tanpa_nama = jasa.loc[jasa['TEKNISI'] == 'TIDAK ADA TEKNISI', 'TOTAL HARGA'].sum()

n_kw = (jasa['TARIF_LABEL'] != LABEL_LAINNYA).sum()
if n_kw == 0:
    sama = abs(tarif_lain - tarif_flat) < 1e-9
    st.warning(
        "Pada periode ini **tidak ada item jasa yang mengandung kata kunci** "
        "(Interface / Normal / Mati Total / Promo) — semuanya memakai penamaan lama "
        f"seperti `JASA REPAIR`, sehingga kena tarif {tarif_lain:.0f}%"
        + (f", dan karena pembanding juga {tarif_flat:.0f}% kedua skema jadi **sama persis**."
           if sama else ".")
        + " Penamaan berkata kunci baru mulai dipakai sekitar Juli 2026.")
elif n_kw < len(jasa) * 0.5:
    st.info(f"Baru **{n_kw:,} dari {len(jasa):,} baris** ({n_kw/len(jasa)*100:.0f}%) "
            "memakai penamaan berkata kunci; sisanya kena tarif tanpa-kata-kunci.")

st.markdown(kpi_html([
    {'label': 'Omzet Jasa', 'value': rp(omzet), 'sub': f"{len(jasa):,} baris",
     'grad': 'linear-gradient(135deg,#1f3864,#2e5394)'},
    {'label': 'Bagi Hasil (Aturan)', 'value': rp(bh),
     'sub': f"{(bh/omzet*100 if omzet else 0):.1f}% dari omzet jasa",
     'grad': 'linear-gradient(135deg,#16a34a,#22c55e)'},
    {'label': f'Pembanding Flat {tarif_flat:.0f}%', 'value': rp(fl),
     'sub': f'omzet jasa × {tarif_flat:.0f}%',
     'grad': 'linear-gradient(135deg,#7c3aed,#a855f7)'},
    {'label': 'Selisih', 'value': rp(selisih),
     'sub': ('aturan lebih besar' if selisih > 0
             else 'flat lebih besar' if selisih < 0 else 'sama'),
     'grad': ('linear-gradient(135deg,#e0921f,#e2b21a)' if selisih >= 0
              else 'linear-gradient(135deg,#c9392f,#e0475a)')},
    {'label': 'Jumlah Teknisi', 'value': f"{n_tek:,}",
     'sub': f"rata-rata {rp(bh/n_tek if n_tek else 0)}/teknisi",
     'grad': 'linear-gradient(135deg,#0f8a82,#17a3a3)'},
    {'label': 'Omzet Tanpa Nama Teknisi', 'value': rp(tanpa_nama),
     'sub': f"{(jasa['TEKNISI'] == 'TIDAK ADA TEKNISI').sum():,} baris",
     'grad': 'linear-gradient(135deg,#64748b,#94a3b8)'},
]), unsafe_allow_html=True)
st.write("")

lbl_flat = f'Pembanding {tarif_flat:.0f}%'

# ---------------------------------------------------------------------------
# Rekap utama: per Teknisi x Cabang
# ---------------------------------------------------------------------------
st.markdown("### Rekap Bagi Hasil per Teknisi & Cabang")
st.caption(
    "Dipecah per cabang karena sebagian teknisi bekerja di lebih dari satu cabang, "
    "sehingga bagi hasilnya bisa dibebankan ke cabang yang tepat."
)

rek = (jasa_tampil.groupby(['TEKNISI', 'CABANG'], as_index=False)
       .agg(Baris=('TOTAL HARGA', 'size'),
            Omzet_Jasa=('TOTAL HARGA', 'sum'),
            Bagi_Hasil=('BAGI_HASIL', 'sum'),
            Flat=('FLAT', 'sum')))
rek['Selisih'] = rek['Bagi_Hasil'] - rek['Flat']
rek['Efektif %'] = (rek['Bagi_Hasil'] / rek['Omzet_Jasa'] * 100).round(1)
rek = rek.sort_values('Bagi_Hasil', ascending=False)

rek_show = rek.rename(columns={
    'TEKNISI': 'Nama Teknisi', 'CABANG': 'Cabang',
    'Omzet_Jasa': 'Omzet Jasa', 'Bagi_Hasil': 'Bagi Hasil (Aturan)', 'Flat': lbl_flat})

cari = st.text_input("Cari nama teknisi / cabang", key='cari_rekap')
rek_view = rek_show
if cari:
    m = rek_show.apply(lambda r: cari.upper() in
                       f"{r['Nama Teknisi']} {r['Cabang']}".upper(), axis=1)
    rek_view = rek_show[m]

st.dataframe(
    rek_view.style.format({
        'Baris': '{:,.0f}', 'Omzet Jasa': 'Rp {:,.0f}',
        'Bagi Hasil (Aturan)': 'Rp {:,.0f}', lbl_flat: 'Rp {:,.0f}',
        'Selisih': 'Rp {:,.0f}'}),
    use_container_width=True, height=460, hide_index=True, key='tabel_rekap')

# --- unduhan: wajib memuat Nama Teknisi, Cabang, Bagi Hasil (Aturan) ---
unduh = rek_show[['Nama Teknisi', 'Cabang', 'Bagi Hasil (Aturan)',
                  'Omzet Jasa', lbl_flat, 'Selisih', 'Baris', 'Efektif %']].copy()
for c in ['Bagi Hasil (Aturan)', 'Omzet Jasa', lbl_flat, 'Selisih']:
    unduh[c] = unduh[c].round(0).astype('int64')

u1, u2 = st.columns(2)
with u1:
    st.download_button(
        "⬇️ Unduh rekap per Teknisi & Cabang (CSV)",
        data=unduh.to_csv(index=False).encode('utf-8-sig'),
        file_name=f"bagi_hasil_teknisi_cabang_{tag_file}.csv",
        mime="text/csv", use_container_width=True, key='unduh_rekap')
with u2:
    gab = (jasa_tampil.groupby('TEKNISI', as_index=False)
           .agg(Omzet_Jasa=('TOTAL HARGA', 'sum'),
                Bagi_Hasil=('BAGI_HASIL', 'sum'),
                Flat=('FLAT', 'sum')))
    gab['Cabang'] = gab['TEKNISI'].map(
        jasa_tampil.groupby('TEKNISI')['CABANG']
        .apply(lambda s: ', '.join(sorted(s.unique()))))
    gab['Selisih'] = gab['Bagi_Hasil'] - gab['Flat']
    gab = gab.rename(columns={'TEKNISI': 'Nama Teknisi', 'Omzet_Jasa': 'Omzet Jasa',
                              'Bagi_Hasil': 'Bagi Hasil (Aturan)', 'Flat': lbl_flat})
    gab = gab[['Nama Teknisi', 'Cabang', 'Bagi Hasil (Aturan)', 'Omzet Jasa',
               lbl_flat, 'Selisih']].sort_values('Bagi Hasil (Aturan)', ascending=False)
    for c in ['Bagi Hasil (Aturan)', 'Omzet Jasa', lbl_flat, 'Selisih']:
        gab[c] = gab[c].round(0).astype('int64')
    st.download_button(
        "⬇️ Unduh rekap per Teknisi (digabung semua cabang)",
        data=gab.to_csv(index=False).encode('utf-8-sig'),
        file_name=f"bagi_hasil_teknisi_{tag_file}.csv",
        mime="text/csv", use_container_width=True, key='unduh_gab')

st.caption("Kedua berkas memuat kolom **Nama Teknisi**, **Cabang**, dan "
           "**Bagi Hasil (Aturan)**, ditambah omzet, pembanding, dan selisihnya.")

# ---------------------------------------------------------------------------
# Unduhan Excel (multi-sheet: rekap + satu sheet per cabang)
# ---------------------------------------------------------------------------
KATEGORI_ORDER = ['Interface', 'Normal', 'Mati Total', 'Promo', LABEL_LAINNYA]


def _sheet_name(nama, terpakai):
    """Nama sheet Excel yang aman: <=31 karakter, tanpa karakter terlarang, unik."""
    s = str(nama)
    for ch in '[]:*?/\\':
        s = s.replace(ch, '-')
    s = s.strip() or 'Cabang'
    s = s[:31]
    dasar, n = s, 2
    while s.lower() in terpakai:
        akhiran = f"_{n}"
        s = dasar[:31 - len(akhiran)] + akhiran
        n += 1
    terpakai.add(s.lower())
    return s


def rekap_kualifikasi(df, keys):
    """Rekap omzet & bagi hasil, dipecah per kualifikasi (Interface/Normal/Mati Total/...)."""
    base = (df.groupby(keys, as_index=False)
              .agg(Baris=('TOTAL HARGA', 'size'),
                   Omzet=('TOTAL HARGA', 'sum'),
                   BH=('BAGI_HASIL', 'sum'),
                   Flat=('FLAT', 'sum')))

    def _pivot(nilai, prefix):
        p = df.pivot_table(index=keys, columns='TARIF_LABEL', values=nilai,
                           aggfunc='sum', fill_value=0.0)
        for k in KATEGORI_ORDER:
            if k not in p.columns:
                p[k] = 0.0
        p = p[KATEGORI_ORDER]
        p.columns = [f"{prefix} {k}" for k in KATEGORI_ORDER]
        return p.reset_index()

    out = (base.merge(_pivot('TOTAL HARGA', 'Omzet'), on=keys, how='left')
               .merge(_pivot('BAGI_HASIL', 'Bagi Hasil'), on=keys, how='left'))
    out['Selisih'] = out['BH'] - out['Flat']
    out['Efektif %'] = (out['BH'] / out['Omzet'].replace(0, pd.NA) * 100).round(1)

    urut = list(keys) + ['Baris'] \
        + [f"Omzet {k}" for k in KATEGORI_ORDER] + ['Omzet'] \
        + [f"Bagi Hasil {k}" for k in KATEGORI_ORDER] + ['BH', 'Flat', 'Selisih', 'Efektif %']
    out = out[urut].rename(columns={
        'TEKNISI': 'Nama Teknisi', 'CABANG': 'Cabang',
        'Omzet': 'Omzet Jasa (Total)', 'BH': 'Bagi Hasil (Aturan)', 'Flat': lbl_flat})
    return out.sort_values('Bagi Hasil (Aturan)', ascending=False).reset_index(drop=True)


def _tulis_sheet(writer, df, nama_sheet, judul):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    df.to_excel(writer, sheet_name=nama_sheet, index=False, startrow=3)
    ws = writer.sheets[nama_sheet]

    ws.cell(row=1, column=1, value=judul).font = Font(bold=True, size=13, color='1F3864')
    ws.cell(row=2, column=1,
            value=f"Periode: {periode_txt} · Tarif: "
                  + ", ".join(f"{k} {v:.0f}%" for k, v in tarif_input.items())
                  + f", Lainnya {tarif_lain:.0f}%, pembanding flat {tarif_flat:.0f}%"
            ).font = Font(size=9, italic=True, color='555555')

    n_baris, n_kol = len(df), len(df.columns)
    head_fill = PatternFill('solid', fgColor='1F3864')
    head_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='D9D9D9')

    for j, kol in enumerate(df.columns, start=1):
        c = ws.cell(row=4, column=j)
        c.fill, c.font = head_fill, head_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(bottom=Side(style='medium', color='1F3864'))
        lebar = max(len(str(kol)) + 2,
                    (df[kol].astype(str).str.len().max() if n_baris else 0) + 2)
        ws.column_dimensions[get_column_letter(j)].width = min(max(lebar, 10), 34)

    kol_rp = [c for c in df.columns
              if c.startswith(('Omzet', 'Bagi Hasil')) or c in (lbl_flat, 'Selisih')]
    idx_rp = [df.columns.get_loc(c) + 1 for c in kol_rp]
    idx_baris = (df.columns.get_loc('Baris') + 1) if 'Baris' in df.columns else None
    idx_pct = (df.columns.get_loc('Efektif %') + 1) if 'Efektif %' in df.columns else None

    for i in range(n_baris):
        r = 5 + i
        if i % 2 == 1:
            for j in range(1, n_kol + 1):
                ws.cell(row=r, column=j).fill = PatternFill('solid', fgColor='F4F7FB')
        for j in idx_rp:
            ws.cell(row=r, column=j).number_format = '#,##0'
        if idx_baris:
            ws.cell(row=r, column=idx_baris).number_format = '#,##0'
        if idx_pct:
            ws.cell(row=r, column=idx_pct).number_format = '0.0'
        for j in range(1, n_kol + 1):
            ws.cell(row=r, column=j).border = Border(bottom=thin)

    # baris TOTAL
    if n_baris:
        rt = 5 + n_baris
        ws.cell(row=rt, column=1, value='TOTAL')
        for j in range(1, n_kol + 1):
            c = ws.cell(row=rt, column=j)
            c.font = Font(bold=True)
            c.fill = PatternFill('solid', fgColor='DCE6F1')
            c.border = Border(top=Side(style='medium', color='1F3864'))
        for j in idx_rp + ([idx_baris] if idx_baris else []):
            L = get_column_letter(j)
            c = ws.cell(row=rt, column=j, value=f"=SUM({L}5:{L}{rt-1})")
            c.number_format = '#,##0'
            c.font = Font(bold=True)
        if idx_pct and 'Bagi Hasil (Aturan)' in df.columns:
            Lb = get_column_letter(df.columns.get_loc('Bagi Hasil (Aturan)') + 1)
            Lo = get_column_letter(df.columns.get_loc('Omzet Jasa (Total)') + 1)
            c = ws.cell(row=rt, column=idx_pct,
                        value=f"=IF({Lo}{rt}=0,0,{Lb}{rt}/{Lo}{rt}*100)")
            c.number_format = '0.0'
            c.font = Font(bold=True)

    ws.freeze_panes = ws.cell(row=5, column=1)
    if n_baris:
        ws.auto_filter.ref = f"A4:{get_column_letter(n_kol)}{4 + n_baris}"


def buat_excel(df_sumber):
    """Workbook: Ringkasan + Rekap Teknisi & Cabang + Rekap per Cabang + sheet per cabang."""
    buf = io.BytesIO()

    # normalisasi kunci supaya tidak ada baris yang hilang saat groupby
    d = df_sumber.copy()
    d['TEKNISI'] = (d['TEKNISI'].fillna('TIDAK ADA TEKNISI').astype(str).str.strip()
                    .replace({'': 'TIDAK ADA TEKNISI', 'nan': 'TIDAK ADA TEKNISI',
                              'NaN': 'TIDAK ADA TEKNISI', 'None': 'TIDAK ADA TEKNISI'}))
    d['CABANG'] = (d['CABANG'].fillna('(TANPA CABANG)').astype(str).str.strip()
                   .replace({'': '(TANPA CABANG)', 'nan': '(TANPA CABANG)'}))

    rek_all = rekap_kualifikasi(d, ['TEKNISI', 'CABANG'])
    rek_cab = rekap_kualifikasi(d, ['CABANG'])
    rek_tek = rekap_kualifikasi(d, ['TEKNISI'])

    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        _tulis_sheet(writer, rek_all, 'Rekap Teknisi & Cabang',
                     'Rekap Bagi Hasil per Teknisi & Cabang')
        _tulis_sheet(writer, rek_tek, 'Rekap Teknisi',
                     'Rekap Bagi Hasil per Teknisi (gabungan semua cabang)')
        _tulis_sheet(writer, rek_cab, 'Rekap Cabang', 'Rekap Bagi Hasil per Cabang')

        terpakai = {'rekap teknisi & cabang', 'rekap teknisi', 'rekap cabang'}
        for cab in sorted(d['CABANG'].unique()):
            sub = d[d['CABANG'] == cab]
            if sub.empty:
                continue
            dc = rekap_kualifikasi(sub, ['TEKNISI']).copy()
            dc.insert(1, 'Cabang', cab)
            _tulis_sheet(writer, dc, _sheet_name(cab, terpakai),
                         f'Bagi Hasil Teknisi — Cabang {cab}')
    buf.seek(0)
    return buf.getvalue()


with st.container():
    st.markdown("##### 📊 Unduh Excel (multi-sheet per cabang)")
    st.caption(
        "Berisi kolom **Nama Teknisi**, **Cabang**, **Omzet**, rincian kualifikasi "
        "**Interface / Normal / Mati Total / Promo / Lainnya** (omzet & bagi hasil "
        "masing-masing), serta **Bagi Hasil**. Sheet: rekap gabungan, rekap per teknisi, "
        "rekap per cabang, lalu satu sheet untuk tiap cabang."
    )
    if st.button("🧾 Siapkan berkas Excel", key='siap_xlsx', use_container_width=True):
        with st.spinner("Menyusun workbook..."):
            st.session_state['xlsx_bytes'] = buat_excel(jasa_tampil)
            st.session_state['xlsx_tag'] = tag_file
    if st.session_state.get('xlsx_bytes') is not None:
        st.download_button(
            "⬇️ Unduh Excel (.xlsx)",
            data=st.session_state['xlsx_bytes'],
            file_name=f"bagi_hasil_teknisi_{st.session_state.get('xlsx_tag', tag_file)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key='unduh_xlsx')



# ---------------------------------------------------------------------------
# Grafik & rekap pendukung
# ---------------------------------------------------------------------------
g1, g2 = st.columns([1.15, 1])
with g1:
    st.markdown("#### 15 Teratas — Aturan vs Pembanding")
    top = rek[rek['TEKNISI'] != 'TIDAK ADA TEKNISI'].head(15).copy()
    top['NAMA'] = top['TEKNISI'].str.slice(0, 22) + " — " + top['CABANG'].str.slice(0, 10)
    top = top.sort_values('Bagi_Hasil')
    fig = go.Figure()
    fig.add_bar(y=top['NAMA'], x=top['Bagi_Hasil'], orientation='h',
                name='Aturan', marker_color='#16a34a')
    fig.add_bar(y=top['NAMA'], x=top['Flat'], orientation='h',
                name=f'Flat {tarif_flat:.0f}%', marker_color='#a855f7')
    fig.update_layout(barmode='group', height=560, margin=dict(l=10, r=10, t=10, b=10),
                      legend=dict(orientation='h', y=1.04), xaxis_title='Rupiah')
    st.plotly_chart(fig, use_container_width=True, key='fig_top')

with g2:
    st.markdown("#### Komposisi Omzet Jasa per Tarif")
    gtar = jasa.groupby('TARIF_LABEL').agg(
        Baris=('TOTAL HARGA', 'size'), Omzet=('TOTAL HARGA', 'sum'),
        Bagi_Hasil=('BAGI_HASIL', 'sum'))
    gtar['Tarif'] = gtar.index.map(lambda k: peta_tarif.get(k, 0.0) * 100)
    gtar['Tarif'] = gtar['Tarif'].round(1).astype(str) + '%'
    gtar = gtar.sort_values('Omzet', ascending=False)
    st.dataframe(
        gtar[['Tarif', 'Baris', 'Omzet', 'Bagi_Hasil']]
        .rename(columns={'Bagi_Hasil': 'Bagi Hasil'})
        .style.format({'Baris': '{:,.0f}', 'Omzet': 'Rp {:,.0f}',
                       'Bagi Hasil': 'Rp {:,.0f}'}),
        use_container_width=True, key='tabel_tarif')
    figp = px.pie(names=gtar.index, values=gtar['Omzet'], hole=0.55,
                  color_discrete_sequence=PALETTE)
    figp.update_layout(height=300, margin=dict(l=5, r=5, t=5, b=5),
                       legend=dict(font=dict(size=9)))
    st.plotly_chart(figp, use_container_width=True, key='fig_tarif')

st.markdown("#### Rekap per Cabang")
gcb = jasa.groupby('CABANG', as_index=False).agg(
    Teknisi=('TEKNISI', 'nunique'), Baris=('TOTAL HARGA', 'size'),
    Omzet_Jasa=('TOTAL HARGA', 'sum'), Bagi_Hasil=('BAGI_HASIL', 'sum'),
    Flat=('FLAT', 'sum'))
gcb['Selisih'] = gcb['Bagi_Hasil'] - gcb['Flat']
gcb['Efektif %'] = (gcb['Bagi_Hasil'] / gcb['Omzet_Jasa'] * 100).round(1)
gcb = gcb.sort_values('Bagi_Hasil', ascending=False).rename(columns={
    'CABANG': 'Cabang', 'Omzet_Jasa': 'Omzet Jasa',
    'Bagi_Hasil': 'Bagi Hasil (Aturan)', 'Flat': lbl_flat})
st.dataframe(
    gcb.style.format({'Teknisi': '{:,.0f}', 'Baris': '{:,.0f}',
                      'Omzet Jasa': 'Rp {:,.0f}', 'Bagi Hasil (Aturan)': 'Rp {:,.0f}',
                      lbl_flat: 'Rp {:,.0f}', 'Selisih': 'Rp {:,.0f}'}),
    use_container_width=True, height=380, hide_index=True, key='tabel_cabang')
st.download_button(
    "⬇️ Unduh rekap per Cabang (CSV)",
    data=gcb.to_csv(index=False).encode('utf-8-sig'),
    file_name=f"bagi_hasil_cabang_{tag_file}.csv", mime="text/csv", key='unduh_cab')

st.markdown("#### Detail Transaksi Jasa")
q = st.text_input("Cari teknisi / cabang / barang / faktur", key='cari_detail')
kol = ['TGL FAKTUR', 'NO FAKTUR', 'CABANG', 'TEKNISI', 'NAMA BARANG',
       'TARIF_LABEL', 'TARIF', 'TOTAL HARGA', 'BAGI_HASIL', 'FLAT']
kol = [c for c in kol if c in jasa.columns]
det = jasa[kol].rename(columns={
    'CABANG': 'Cabang', 'TEKNISI': 'Nama Teknisi', 'TARIF_LABEL': 'Kategori Tarif',
    'TARIF': 'Tarif', 'BAGI_HASIL': 'Bagi Hasil', 'FLAT': lbl_flat})
if q:
    m = det.apply(lambda r: q.upper() in ' '.join(str(v) for v in r.values).upper(), axis=1)
    det = det[m]
st.caption(f"{len(det):,} baris (ditampilkan maksimal 1.000).")
st.dataframe(det.head(1000), use_container_width=True, height=360,
             hide_index=True, key='tabel_detail')

with st.expander("ℹ️ Cara perhitungan & catatan"):
    st.write(
        "**Tarif bagi hasil** ditentukan dari kata kunci pada kolom NAMA BARANG, "
        "mengikuti isian pada panel Pengaturan Tarif di atas:\n"
        f"- mengandung **Interface** → {tarif_input['Interface']:.0f}%\n"
        f"- mengandung **Normal** → {tarif_input['Normal']:.0f}%\n"
        f"- mengandung **Mati Total** → {tarif_input['Mati Total']:.0f}%\n"
        f"- mengandung **Promo** → {tarif_input['Promo']:.0f}%\n"
        f"- tanpa kata kunci mana pun → **{tarif_lain:.0f}%** (mencakup item berpola "
        "`JASA ...` seperti JASA REPAIR, JASA BATERAI, JASA LCD 50%)\n\n"
        "Bila satu nama mengandung dua kata kunci sekaligus (mis. "
        f"`JS PROMO LCD 250K - NORMAL`), dipakai **{prioritas} "
        f"{tarif_input[prioritas]:.0f}%** sesuai pilihan prioritas.\n\n"
        "**Periode penggajian** memakai cutoff tanggal 24 s/d 23: gaji bulan M dihitung "
        "dari 24 bulan (M−2) sampai 23 bulan (M−1). Contoh gaji Mei 2026 = 24 Maret 2026 "
        "s/d 23 April 2026. Tanggal acuan: **TGL FAKTUR**.\n\n"
        f"**Pembanding Flat {tarif_flat:.0f}%** = seluruh omzet jasa × {tarif_flat:.0f}%, "
        "tanpa membedakan jenis pekerjaan.\n\n"
        "Nama teknisi diambil dari kolom **NAMA TEKNISI (FINAL)**; bila kosong dipakai "
        "kolom NAMA TEKNISI. Baris yang keduanya kosong masuk kelompok "
        "*TIDAK ADA TEKNISI* — tetap ditampilkan agar terlihat, dan bisa disembunyikan "
        "lewat centang di atas.\n\n"
        "Perhitungan memakai **omzet jasa (TOTAL HARGA)**, belum dikurangi biaya apa pun. "
        "Hanya baris berkategori **JASA** yang dihitung."
    )
